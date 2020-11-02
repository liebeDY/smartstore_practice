import requests
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook

# 사이트 주소 설정/1장 시작
baseurl = "https://www.bibleserver.com/SLT/"
Biblename = input('원하는 성경을 입력하세요(독일어) :')
biblekapital = 1
endkapital = input('끝나는 장을 입력하세요(숫자) :')


wb = openpyxl.Workbook()
for wieder in range(int(biblekapital), int(endkapital)+1):
    
    url = baseurl + quote_plus(Biblename+str(wieder))
    response = requests.get(url)

    soup = BeautifulSoup(response.content, "html.parser")

    bibletext = soup.select("span.verse-content--hover")
    

    sheet = wb.create_sheet(title = f'{wieder}장',index=-1)
    for text in bibletext:
        t = []
        text = text.get_text()
        t.append(text)

        sheet.append([text])
        


print('완료되었습니다')

wb.save(f'{Biblename}.xlsx')

