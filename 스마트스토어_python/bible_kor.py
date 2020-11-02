import requests
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook

biblenumber = input('원하는 성경의 순서를 입력하세요(숫자) :')
biblename = input('성경 이름 :')

startkapital = 1
endkapital = input('끝나는 장을 입력하세요(숫자) :')

baseurl = "http://www.holybible.or.kr/mobile/B_GAE/cgi-m/bibleftxt.php?VR=GAE&"
biblename_url = f"VL={quote_plus(biblenumber)}&"

wb = openpyxl.Workbook()

for wieder in range(int(startkapital), int(endkapital)+1):
    kapital_url = "CN="+str(wieder)+"&"
    url = baseurl + biblename_url + kapital_url + "CV=99"
    response = requests.get(url)
    
    soup = BeautifulSoup(response.content, "html.parser")

    bibletext = soup.select('font.tk4l')
    sheet = wb.create_sheet(title = f'{wieder}장',index=-1)

    for text in bibletext:
        t = []
        text = text.get_text()
        t.append(text)

        sheet.append([text])


    print(str(wieder)+"장 완료")
print()
print(f'{biblenumber}, {biblename} 완료되었습니다')

wb.save(f'{biblename}.xlsx')