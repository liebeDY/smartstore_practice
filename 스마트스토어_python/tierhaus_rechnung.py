from openpyxl import load_workbook

wb = load_workbook("C:/Users/Dong/Downloads/스마트스토어_선택주문조회_20200923_1142.xlsx", data_only=True)

sheet = wb['발주발송관리']

def column_return(keyword):
    information_dict = {
        "주문번호" : 2,
        "송장번호" : 6,
        "구매자명" : 9,
        "수취인명" : 11,
        "결제일" : 15,
        "상품명" : 17,
        "옵션정보" : 19,
        "수량" : 21,
        "상품가격" : 23,
        "상품별총주문금액" : 26,
        "배송비합계" : 35,
        "제주도서배송비추가" : 36,
        "수취인연락처" : 41,
        "기본주소" : 64,
        "상세주소" : 65
    }
    index = information_dict.get(keyword)
    
    return sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=index, max_col=index)

def valuelist_return(variable):
    value_list = []
    for values in variable:
        for value in values:
            value_list.append(value.value)
    return value_list

def dict_maker(orderlist,invoicelist,kundenlist,empaengerlist,paymentlist,productlist,optionlist,quantitylist,pricelist,totalpricelist,deliverylist,jejulist,telephonlist,adresselist,adresse2list):
    orderdict_list = []
    for order,invoice,kunden,empaenger,payment,product,option,quantity,price,totalprice,delivery,jeju,telephon,adresse,adresse2 in zip(orderlist,invoicelist,kundenlist,empaengerlist,paymentlist,productlist,optionlist,quantitylist,pricelist,totalpricelist,deliverylist,jejulist,telephonlist,adresselist,adresse2list):
        print(price)
        order_dict = {
            "주문번호" : order,
            "송장번호" : invoice,
            "구매자명" : kunden,
            "수취인명" : empaenger,
            "결제일" : payment,
            "상품명" : product,
            "옵션정보" : option,
            "수량" : int(quantity),
            "상품가격" : int(price),
            "상품별총주문금액" : int(totalprice),
            "배송비합계" : int(delivery),
            "제주도서배송비추가" : int(jeju),
            "수취인연락처" : telephon,
            "기본주소" : adresse,
            "상세주소" : adresse2
        }
        print(order_dict)
        orderdict_list.append(order_dict)
    return orderdict_list


order_number_rng = column_return("주문번호")
invoice_number_rng = column_return("송장번호")
kunden_name_rng = column_return("구매자명")
empaenger_name_rng = column_return("수취인명")
payment_date_rng = column_return("결제일")
product_name_rng = column_return("상품명")
option_info_rng = column_return("옵션정보")
quantity_rng = column_return("수량")
product_price_rng = column_return("상품가격")
product_totalprice_rng = column_return("상품별총주문금액")
delivery_totalprice_rng = column_return("배송비합계")
delivery_jeju_price_rng = column_return("제주도서배송비추가")
telephon_rng = column_return("수취인연락처")
adresse_rng = column_return("기본주소")
adresse2_rng = column_return("상세주소")

order_number_list = valuelist_return(order_number_rng)
invoice_number_list = valuelist_return(invoice_number_rng)
kunden_name_list = valuelist_return(kunden_name_rng)
empaenger_name_list = valuelist_return(empaenger_name_rng)
payment_date_list = valuelist_return(payment_date_rng)
product_name_list = valuelist_return(product_name_rng)
option_info_list = valuelist_return(option_info_rng)
quantity_list = valuelist_return(quantity_rng)
product_price_list = valuelist_return(product_price_rng)
product_totalprice_list = valuelist_return(product_totalprice_rng)
delivery_totalprice_list = valuelist_return(delivery_totalprice_rng)
delivery_jeju_price_list = valuelist_return(delivery_jeju_price_rng)
telephon_list = valuelist_return(telephon_rng)
adresse_list = valuelist_return(adresse_rng)
adresse2_list = valuelist_return(adresse2_rng)

orderdict_a = dict_maker(order_number_list,invoice_number_list,kunden_name_list,empaenger_name_list,payment_date_list,product_name_list,option_info_list,quantity_list,product_price_list,product_totalprice_list,delivery_totalprice_list,delivery_jeju_price_list,telephon_list,adresse_list,adresse2_list)
print(orderdict_a)
rechnung_wb = load_workbook("C:/Users/Dong/Desktop/테스트.xlsx", data_only=True)
rechnung_ws = rechnung_wb.active
for order in orderdict_a:
    order = list(order.values())
    print(order)
    rechnung_ws.append(order)
    # rechnung_ws.cell(1,1,order["수취인명"])
    # rechnung_ws.cell(1,2,order["기본주소"])
    # rechnung_ws.cell(1,3,order["상세주소"])
    # rechnung_ws.cell(1,4,order["수취인연락처"])


rechnung_wb.save("C:/Users/Dong/Desktop/테스트3.xlsx")


# idx = 1
# for i in orderdict_a:
#     print(i.get('구매자명'))
#     rechnung_ws.cell(9,1,i.get('구매자명'))
#     rechnung_wb.save(f"C:/Users/Dong/Desktop/테스트{str(idx)}.xlsx")
#     idx += 1